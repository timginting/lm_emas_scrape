from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
SELL_FILE = DATA_DIR / "antam_sell.json"
BUYBACK_FILE = DATA_DIR / "antam_buyback.json"
OUTPUT_FILE = DATA_DIR / "harga_emas_antam.xlsx"
JAKARTA = ZoneInfo("Asia/Jakarta")


def load_records(path: Path) -> list[list[int]]:
    with path.open("r", encoding="utf-8") as file:
        values = json.load(file)
    return [
        [int(item[0]), int(item[1])]
        for item in values
        if isinstance(item, list) and len(item) == 2 and item[0] is not None and item[1] is not None
    ]


def local_date(timestamp_ms: int):
    return datetime.fromtimestamp(timestamp_ms / 1000, tz=JAKARTA).date()


def build_workbook() -> int:
    by_date: dict = {}
    for timestamp, price in load_records(SELL_FILE):
        by_date.setdefault(local_date(timestamp), {})["sell"] = price
    for timestamp, price in load_records(BUYBACK_FILE):
        by_date.setdefault(local_date(timestamp), {})["buyback"] = price

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Harga Emas"
    headers = [
        "Tanggal",
        "Harga Jual ANTAM (Rp/gram)",
        "Harga Buyback (Rp/gram)",
        "Spread (Rp/gram)",
    ]
    sheet.append(headers)

    for day in sorted(by_date):
        sell = by_date[day].get("sell")
        buyback = by_date[day].get("buyback")
        spread = sell - buyback if sell is not None and buyback is not None else None
        sheet.append([day, sell, buyback, spread])

    header_fill = PatternFill("solid", fgColor="B58A1F")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 14
    for column in range(2, 5):
        sheet.column_dimensions[get_column_letter(column)].width = 30
        for cell in sheet[get_column_letter(column)][1:]:
            cell.number_format = 'Rp #,##0;[Red]-Rp #,##0'
    for cell in sheet["A"][1:]:
        cell.number_format = "dd/mm/yyyy"

    metadata = workbook.create_sheet("Metadata")
    metadata.append(["Keterangan", "Nilai"])
    metadata.append(["Sumber data", "https://www.logammulia.com/id/grafik-harga-emas"])
    metadata.append(["Diperbarui (Asia/Jakarta)", datetime.now(JAKARTA).strftime("%Y-%m-%d %H:%M:%S %Z")])
    metadata.append(["Jumlah baris", len(by_date)])
    metadata.column_dimensions["A"].width = 30
    metadata.column_dimensions["B"].width = 65
    for cell in metadata[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_FILE)
    return len(by_date)


if __name__ == "__main__":
    count = build_workbook()
    print(f"Berhasil membuat {OUTPUT_FILE} ({count} baris).")

